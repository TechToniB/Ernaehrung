from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Basisverzeichnis festlegen (z.B. als relativer Pfad zum Skript)
base_dir = Path.home() / "Dokumente" / "Ernaehrung" / "Gesunde Ernährung"

# Excel Datei einlesen
excel_path = base_dir / "Quellen" / "DGE Ist Wundervoll" / "GibMirAlles.xlsx"
df = pd.read_excel(excel_path)

# Daten filtern
filtered_Y = df[
    (df['Bevölkerungsgruppe'] == '25 bis unter 51 Jahre') &
    (df['Geschlecht'] == 'Männlich')
]
filtered_A = df[
    (df['Bevölkerungsgruppe'] == '25 bis unter 51 Jahre') &
    (df['Geschlecht'] == 'Weiblich')
]

# Ausgabeordner
output_dir = base_dir / "Filter"
output_dir.mkdir(exist_ok=True)

def einfärben(path):
    wb = load_workbook(path)
    ws = wb.active

    # Spaltenindex für "Einheit" und "Kategorie" finden
    einheit_col = None
    kategorie_col = None
    for idx, cell in enumerate(ws[1], 1):
        header = str(cell.value).strip().lower()
        if header == "einheit":
            einheit_col = idx
        if header == "kategorie":
            kategorie_col = idx

    # Farben für verschiedene Einheiten
    farben_einheit = {
        "mg/tag": "FFFF99",           # Gelb
        "µg/tag": "CCFFCC",           # Grün
        "g/tag": "FFCCCC",            # Rot
        "kcal/tag": "CCE5FF",         # Blau
        "% der energie": "FFD966",    # Orange
    }
    # Farben für verschiedene Kategorie-Werte
    farben_kategorie = {
        "richtwert": "FFF2CC",            # Hellgelb
        "schätzwert": "D9EAD3",           # Hellgrün
        "empfohlene zufuhr": "CFE2F3",    # Hellblau
        # Weitere Werte/Farben nach Bedarf ergänzen
    }

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
        # Einheit einfärben
        if einheit_col:
            einheit = str(row[einheit_col-1].value).strip().lower()
            if einheit in farben_einheit:
                row[einheit_col-1].fill = PatternFill(start_color=farben_einheit[einheit], end_color=farben_einheit[einheit], fill_type="solid")
        # Kategorie einfärben
        if kategorie_col:
            kategorie = str(row[kategorie_col-1].value).strip().lower()
            if kategorie in farben_kategorie:
                row[kategorie_col-1].fill = PatternFill(start_color=farben_kategorie[kategorie], end_color=farben_kategorie[kategorie], fill_type="solid")

    # Spaltenbreite automatisch anpassen
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # +2 für etwas Abstand

    wb.save(path)

# Gefilterte Dateien speichern (ohne "Alkohol")
yves_path = output_dir / "Yves_Filter.xlsx"
antonia_path = output_dir / "Antonia_Filter.xlsx"

filtered_Y[filtered_Y['Nährstoff'] != 'Alkohol'][['Nährstoff', 'Referenzwert', 'Einheit', 'Kategorie']].to_excel(
    yves_path, index=False)
filtered_A[filtered_A['Nährstoff'] != 'Alkohol'][['Nährstoff', 'Referenzwert', 'Einheit', 'Kategorie']].to_excel(
    antonia_path, index=False)

# Spalte "Einheit" einfärben
einfärben(yves_path)
einfärben(antonia_path)
