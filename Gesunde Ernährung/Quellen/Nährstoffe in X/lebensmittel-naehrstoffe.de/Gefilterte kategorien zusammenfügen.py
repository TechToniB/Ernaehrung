import pandas as pd
import os
from openpyxl import load_workbook

# Pfad zur Excel-Datei
excel_path = os.path.join(os.path.dirname(__file__), 'Gefilterte_Kategorien.xlsx')

# Excel-Datei einlesen
df = pd.read_excel(excel_path)

# Funktion zum Zusammenführen der Werte (außer 'Lebensmittel')
def combine_values(series):
	# Entferne NaN, wandle alles in String, entferne Duplikate, verbinde mit Komma
	return ', '.join(sorted(set(str(x) for x in series if pd.notna(x) and str(x).strip() != '')))

# Gruppieren nach 'Lebensmittel' und alle anderen Spalten zusammenfassen
result = df.groupby('Lebensmittel', as_index=False).agg(combine_values)

# Ergebnis speichern
output_path = os.path.join(os.path.dirname(__file__), 'Gefilterte_Kategorien_zusammengefuegt.xlsx')
result.to_excel(output_path, index=False)

# Spaltenbreite automatisch anpassen
from openpyxl.utils import get_column_letter
wb = load_workbook(output_path)
ws = wb.active
for col in ws.columns:
	max_length = 0
	column = col[0].column  # Spaltenindex (1-basiert)
	for cell in col:
		try:
			if cell.value:
				max_length = max(max_length, len(str(cell.value)))
		except:
			pass
	adjusted_width = max_length + 2
	ws.column_dimensions[get_column_letter(column)].width = adjusted_width
wb.save(output_path)

print(f"Zusammengeführte Tabelle gespeichert als: {output_path}")
